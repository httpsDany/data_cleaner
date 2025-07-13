import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv
import sys

def load_file(filepath):
    ext = Path(filepath).suffix.lower()
    if ext == ".csv":
        with open(filepath, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader)
            expected_cols = len(header)

            for row_number, row in enumerate(reader, start=2):  # start=2 because header is row 1
                if len(row) != expected_cols:
                    # Find first mismatch column index
                    mismatched_cols = [i for i in range(min(len(row), expected_cols)) if ',' in row[i] and not (row[i].startswith('"') and row[i].endswith('"'))]
                    col_hint = f"Possible issue at column {mismatched_cols[0]+1}" if mismatched_cols else "Check currency columns"
                    
                    print(f"\n⚠️  Malformed CSV: Row {row_number} has {len(row)} columns but header has {expected_cols}.")
                    print(f"📍 {col_hint}")
                    print(f"💡 Tip: Check for unquoted currency values with commas")
                    print(f"📌 Fix: Wrap currency values in double quotes to preserve column structure.")
                    exit(1)

        # If valid, load normally
        return pd.read_csv(filepath, quotechar='"', skipinitialspace=True)
    elif ext in [".xlsx", ".xls"]:
        return pd.read_excel(filepath, engine="openpyxl")
    else:
        raise ValueError("Unsupported file type.")


def save_file(df, output_path):
    ext = Path(output_path).suffix.lower()
    if ext == ".csv":
        df.to_csv(output_path, index=False)
    elif ext in [".xlsx", ".xls"]:
        df.to_excel(output_path, index=False)
        if "highlight_nulls" in df.attrs:
            highlight_rows_in_excel(output_path, df.attrs["highlight_nulls"])
    else:
        raise ValueError("Unsupported file type.")

def highlight_rows_in_excel(filepath, row_indices):
    wb = load_workbook(filepath)
    ws = wb.active
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row_idx in row_indices:
        excel_row = row_idx + 2  # +1 for 0-based index, +1 for header
        for col in range(1, ws.max_column + 1):
            ws.cell(row=excel_row, column=col).fill = red_fill

    wb.save(filepath)

